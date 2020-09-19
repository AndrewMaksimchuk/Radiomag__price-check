# Підключаємо бібліотеку для роботи із файлами
from os import listdir
# Бібліотека для створення запросів і отримання відповідей від серевера
import requests
# Бібліотека для роботи із файлами excel
from openpyxl import *
# Список всіх груп товарів з яких отримуємо інформацію в форматі/файлі JSON
import all_stocks
#Бібліотека для запуска зовнішніх програм
import subprocess

# Рядок з якого починаємо записувати дані у новий файл
startRow = 2

# Масив(список) всіх товарів з усіх груп по наясності.
all_goods = []
# Масив(список) всіх товарів з усіх груп де змінилася ціна.
all_goods_with_change_price = []

def create_a_reference_file():
    # Створюємо нову книгу ексель
    wb = Workbook()
    ws = wb.create_sheet("Reference", 0)
    Format = wb["Reference"]
    #Записуємо назви колонок
    Format.cell(row=1, column=1, value="Код товара")
    Format.cell(row=1, column=2, value="Назва товара")
    Format.cell(row=1, column=3, value="Устарівша ціна")
    Format.column_dimensions['A'].width = 11
    Format.column_dimensions['B'].width = 123
    Format.column_dimensions['C'].width = 30
    # Отримуємо всі позиції із груп товарів
    get_all_goods_in_stocks()
    # Заповнюємо таблицю-файл
    filling_in_the_table(Format, startRow)
    # Зберігаємо файл з іменем "New_price"
    file_name = "Reference_table"
    file_extension = ".xlsx"
    wb.save("C:\\Users\\Maksymchuk\\Desktop\\pyPrice\\" + file_name + file_extension)
    print("Еталонна таблиця цін створена!")

def filling_in_the_table(Format, startRow):
        for k in all_goods:
                Format.cell(row=startRow, column=1, value=k.get("code"))
                Format.cell(row=startRow, column=2, value=k.get("name"))
                Format.cell(row=startRow, column=3, value=k.get("price"))
                startRow = startRow + 1

def get_all_goods_in_stocks():
        for i in all_stocks.groups:
                url = 'https://www.rcscomponents.kiev.ua/modules.php?name=Asers_Shop&s_op=productlist_json&group_id=' + str(all_stocks.groups.get(i)) + '&lang=russian'
                request_JSON = requests.get(url)
                # get_array - масив всіх позицій конкретної групи
                get_array = request_JSON.json().get('items')

                for j in get_array:
                        # print(j.get("total_stock"))
                        if (j.get("total_stock")):
                                temp_obj = {
                                        "code": j.get("id"),
                                        "name": j.get("description")[0],
                                        "price": ""
                                        }
                                if (j.get("prices")):
                                        temp_obj["price"] = float(j.get("prices")[0].get("p"))
                                # print(temp_obj)
                                all_goods.append(temp_obj)
        print("Отримано весь товар для звірки цін!")

def checking_prices():
        # Отримуємо всі необхідні позиції товара для звірки цін 
        # і записуємо в глобальний масив(список)
        # як обєкт
        get_all_goods_in_stocks()
        # Рядок з якого починаємо перебирати файл "Reference_table.xlsx"
        start_row = 2
        # Відкриваємо файл "Reference_table.xlsx" для зчитування
        loadWorkbook = load_workbook("C:\\Users\\Maksymchuk\\Desktop\\pyPrice\\" + "Reference_table.xlsx")
        sheet = loadWorkbook.active
        # Максимальна кількість рядків у відкритому файлі
        max_row_range = sheet.max_row

        # Перебираємо всі товари з актуальною ціною і звіряємо з файлом "Reference_table.xlsx"
        for i in all_goods:
                temp_code = i.get("code")
                temp_name = i.get("name")
                temp_price = i.get("price")
                # файл "Reference_table.xlsx"
                for j in range(start_row, max_row_range):
                        code_in_ref_table = sheet.cell(row=j, column=1).value
                        if temp_code == code_in_ref_table:
                                price_in_ref_table = sheet.cell(row=j, column=3).value
                                if temp_price != price_in_ref_table:
                                        sheet.cell(row=j, column=3, value=temp_price)
                                        all_goods_with_change_price.append(i)
                                        continue

        loadWorkbook.save('C:\\Users\\Maksymchuk\\Desktop\\pyPrice\\' + 'Reference_table.xlsx')
        print("Перевірка і оновлення цін закінчена!")
        # print(all_goods_with_change_price)

def table_with_new_price():
        # Флаг для визначення чи були виявлені позиції товара де змінилася ціна. Для інформативності повідомлень
        new_price_is_exist = False
        # Стартовий рядок
        s_row = 2
        # Створюємо нову книгу ексель для позицій із зміненою ціною
        wb_price = Workbook()
        ws_price = wb_price.create_sheet("New_price", 0)
        with_new_price = wb_price["New_price"]

        with_new_price.cell(row=1, column=1, value="Код товара")
        with_new_price.cell(row=1, column=2, value="Назва товара")
        with_new_price.cell(row=1, column=3, value="Актуальна ціна")
        with_new_price.column_dimensions['A'].width = 11
        with_new_price.column_dimensions['B'].width = 123
        with_new_price.column_dimensions['C'].width = 30

        for k in all_goods_with_change_price:
                if k.get("price") != "":
                        with_new_price.cell(row=s_row, column=1, value=k.get("code"))
                        with_new_price.cell(row=s_row, column=2, value=k.get("name"))
                        with_new_price.cell(row=s_row, column=3, value=k.get("price"))
                        s_row = s_row + 1
                        new_price_is_exist = True

        # wb_price.save("C:\\Users\\Maksymchuk\\Desktop\\pyPrice\\" + "New_price.xlsx")
        
        print(new_price_is_exist)
        if new_price_is_exist:
                wb_price.save("C:\\Users\\Maksymchuk\\Desktop\\pyPrice\\" + "New_price.xlsx")
                print("Файл New_price.xlsx оновлено і додано товар у якого змінилася ціна!")
                new_price_is_exist = False
                subprocess.Popen(['C:\\Users\\Maksymchuk\\Desktop\\pyPrice\\phpMail\\php-7.2.7\\php.exe', 'C:\\Users\\Maksymchuk\\Desktop\\pyPrice\\phpMail\\sendMail.php'])
        else:
                print("Файл New_price.xlsx залишився без змін!")

def verify(temp_cell_val, code_in_goods):
        if temp_cell_val == code_in_goods:
                return True
        else:
                return False

def update_reference_table():
        # Відкриваємо еталонну таблицю
        load_book = load_workbook("C:\\Users\\Maksymchuk\\Desktop\\pyPrice\\" + "Reference_table.xlsx")
        sheet_book = load_book.active
        # Максимальна кількість рядків у відкритому файлі
        max_row_range = sheet_book.max_row
        # Рядок з якого починається запис для доповнення таблиці
        up_max_row_range = max_row_range + 1
        # Флаг установки відповіді - чи оновлени еталонну табличку чи ні
        reference_table_is_update = False

        if_i_exist = False
        for i in all_goods:
                code_in_goods = i.get("code")
                
                for j in range(2, max_row_range + 1):
                        temp_cell_val = sheet_book.cell(row=j, column=1).value
                        if verify(temp_cell_val, code_in_goods):
                                if_i_exist = False
                                break
                        else:
                                if_i_exist = True
                # Додаємо новий запис до еталонної таблиці, тому що данної позиції немає в ній. Для майбутньої перевірки ціни
                if if_i_exist == True:
                        sheet_book.cell(row=up_max_row_range, column=1, value=i.get("code"))
                        sheet_book.cell(row=up_max_row_range, column=2, value=i.get("name"))
                        sheet_book.cell(row=up_max_row_range, column=3, value=i.get("price"))
                        up_max_row_range = up_max_row_range + 1
                        if_i_exist = False
                        reference_table_is_update = True
        
        load_book.save("C:\\Users\\Maksymchuk\\Desktop\\pyPrice\\" + "Reference_table.xlsx")
        if reference_table_is_update:
                print("Еталонна таблиця Reference_table.xlsx була оновлена!")
        else:
                print("Еталонна таблиця Reference_table.xlsx залишилася без змін, оновлення не відбулося!")

# Виконання всіх функцій попорядку.
#create_a_reference_file()
print("\n")
print(" Працює скрипт на Python!\n")
checking_prices()
table_with_new_price()
update_reference_table()
print(" Скрипт закінчив роботу.")